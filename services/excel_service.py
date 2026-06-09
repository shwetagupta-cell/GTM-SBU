from openpyxl import load_workbook

from models.entities import EmployeeRecord, ProjectRecord
from services.utils import clean_string, month_label, normalize_emp_code, normalize_name, parse_number


def _load_pandas():
    import pandas as pd

    return pd


def _header_text(value):
    return clean_string(value).lower()


def _find_header_row(dataframe):
    for idx in range(min(len(dataframe.index), 10)):
        values = [_header_text(value) for value in dataframe.iloc[idx].tolist()]
        if any("project name" in value for value in values):
            return idx
    return 0


def _find_col(headers, *phrases):
    for idx, header in enumerate(headers):
        lowered = _header_text(header)
        if any(phrase in lowered for phrase in phrases):
            return idx
    return -1


def _collect_month_columns(headers, start_index):
    month_columns = []
    for idx in range(start_index, len(headers)):
        label = month_label(headers[idx])
        raw = _header_text(headers[idx])
        if raw.startswith("total") or raw.startswith("got td") or "future" in raw or "pv - cf" in raw:
            break
        if raw:
            month_columns.append((label, idx))
    return month_columns


def parse_future_employee_workbook(path):
    pd = _load_pandas()
    workbook = pd.ExcelFile(path)
    employees = []
    for sheet_name in workbook.sheet_names:
        dataframe = pd.read_excel(path, sheet_name=sheet_name, header=None)
        if dataframe.empty:
            continue
        header_row = _find_header_row(dataframe)
        name_row = max(header_row - 1, 0)
        headers = [clean_string(value) for value in dataframe.iloc[header_row].tolist()]
        top_values = [clean_string(value) for value in dataframe.iloc[name_row].tolist()]

        project_name_idx = _find_col(headers, "project name")
        project_id_idx = _find_col(headers, "poject id", "project id")
        project_value_idx = _find_col(headers, "project value")
        future_value_idx = _find_col(headers, "future")
        if project_name_idx < 0 or project_value_idx < 0:
            continue

        cf_months = _collect_month_columns(headers, project_value_idx + 1)
        employee_name_idx = max((idx for idx, value in enumerate(top_values) if value and idx > project_value_idx), default=-1)
        employee_name = top_values[employee_name_idx] if employee_name_idx >= 0 else clean_string(sheet_name)

        incentive_months = []
        if employee_name_idx >= 0:
            for idx in range(employee_name_idx, len(headers)):
                raw = _header_text(headers[idx])
                if raw.startswith("got td") or raw.startswith("total on pv") or "future" in raw:
                    break
                if raw:
                    incentive_months.append((month_label(headers[idx]), idx))

        month_order = []
        for month, _idx in cf_months + incentive_months:
            if month not in month_order:
                month_order.append(month)

        employee = EmployeeRecord(employee_id="", name=employee_name, month_order=month_order)
        for row_index in range(header_row + 1, len(dataframe.index)):
            row = dataframe.iloc[row_index].tolist()
            project_name = clean_string(row[project_name_idx] if project_name_idx < len(row) else "")
            if not project_name or project_name.lower().startswith("total"):
                continue
            project_id = clean_string(row[project_id_idx] if project_id_idx < len(row) else "") or f"{sheet_name}-{row_index}"
            project = ProjectRecord(
                project_id=project_id,
                project_name=project_name,
                project_value=parse_number(row[project_value_idx] if project_value_idx < len(row) else 0),
                future_value=parse_number(row[future_value_idx] if future_value_idx >= 0 and future_value_idx < len(row) else 0),
                source_sheet=sheet_name,
            )
            for month, column_index in cf_months:
                project.monthly_cf[month] = parse_number(row[column_index] if column_index < len(row) else 0)
            for month, column_index in incentive_months:
                project.monthly_incentive[month] = parse_number(row[column_index] if column_index < len(row) else 0)
            employee.projects.append(project)
        employees.append(employee)
    return employees


def parse_team_workbook(path):
    pd = _load_pandas()
    dataframe = pd.read_excel(path, header=None).fillna("")
    rows = dataframe.values.tolist()
    employees = {}
    current_head_id = ""
    current_head_name = ""
    inside_team = False

    for row in rows:
        first_cell = clean_string(row[0] if len(row) > 0 else "")
        second_cell = clean_string(row[1] if len(row) > 1 else "")
        designation = clean_string(row[2] if len(row) > 2 else "")
        location = clean_string(row[3] if len(row) > 3 else "")
        lowered = first_cell.lower()

        if lowered.startswith("head of"):
            current_head_id = ""
            current_head_name = ""
            inside_team = False
            continue
        if lowered.startswith("his team"):
            inside_team = True
            continue
        if lowered.startswith("emp id"):
            continue
        if not first_cell and not second_cell:
            continue

        employee_id = normalize_emp_code(first_cell)
        if not employee_id or not second_cell:
            continue

        record = employees.get(
            employee_id,
            EmployeeRecord(employee_id=employee_id, name=second_cell, designation=designation, location=location),
        )
        record.name = second_cell or record.name
        record.designation = designation or record.designation
        record.location = location or record.location
        employees[employee_id] = record

        if not current_head_id:
            current_head_id = employee_id
            current_head_name = second_cell
            inside_team = False
            continue

        if inside_team and employee_id != current_head_id:
            record.team_head_id = current_head_id
            record.team_head_name = current_head_name

    return employees


def parse_ytd_workbook(path):
    pd = _load_pandas()
    workbook = pd.ExcelFile(path)
    mapping = {}
    for sheet_name in workbook.sheet_names:
        dataframe = pd.read_excel(path, sheet_name=sheet_name)
        for _, row in dataframe.iterrows():
            project_name = clean_string(row.get("Project Name"))
            project_id = clean_string(row.get("Poject ID") or row.get("Project ID"))
            if not project_name and not project_id:
                continue
            meta = {
                "projectValue": parse_number(row.get("Project Value")),
                "totalCf": parse_number(row.get("Total  CF")),
                "pvCf": parse_number(row.get("PV - CF")),
                "receivedPercent": parse_number(row.get("Recived %")),
                "remaining": parse_number(row.get("Remaining")),
                "paid": parse_number(row.get("Paid")),
            }
            mapping[project_id or normalize_name(project_name)] = meta
    return mapping


def parse_closure_workbook(path):
    workbook = load_workbook(path, data_only=True)
    mapping = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        for column_index in range(3, sheet.max_column + 1):
            project_name = clean_string(sheet.cell(2, column_index).value)
            project_id = clean_string(sheet.cell(3, column_index).value)
            if not project_name and not project_id:
                continue

            roles = []
            current_function = ""
            for row_index in range(10, min(sheet.max_row, 31) + 1):
                function_name = clean_string(sheet.cell(row_index, 1).value)
                designation = clean_string(sheet.cell(row_index, 2).value)
                members = clean_string(sheet.cell(row_index, column_index).value)
                if function_name:
                    current_function = function_name
                if designation and members:
                    roles.append(f"{current_function} / {designation} / {members}")

            entry = {
                "projectName": project_name,
                "projectId": project_id,
                "sourcingType": clean_string(sheet.cell(8, column_index).value),
                "roles": roles,
                "sheet": sheet_name,
            }
            if project_id:
                mapping[project_id] = entry
            mapping[normalize_name(project_name)] = entry
    return mapping


def parse_master_workbook(path, closure_map=None, ytd_map=None):
    workbook = load_workbook(path, data_only=True)
    employees = {}

    for sheet_name in workbook.sheetnames:
        if "Team Incentives" not in sheet_name:
            continue
        source_sheet_name = sheet_name.replace(" - Team Incentives", "")
        if source_sheet_name not in workbook.sheetnames:
            continue

        month_name = month_label(source_sheet_name)
        source_sheet = workbook[source_sheet_name]
        team_sheet = workbook[sheet_name]
        project_meta = {}

        for column_index in range(6, source_sheet.max_column + 1, 3):
            value_col = column_index + 1
            project_name = clean_string(source_sheet.cell(2, value_col).value)
            project_id = clean_string(source_sheet.cell(3, value_col).value)
            if not project_name:
                continue
            project_meta[normalize_name(project_name)] = {
                "projectId": project_id or f"{source_sheet_name}-{value_col}",
                "projectName": project_name,
                "projectValue": parse_number(source_sheet.cell(8, value_col).value),
                "cashflow": parse_number(source_sheet.cell(6, value_col).value),
            }

        for row_index in range(2, team_sheet.max_row + 1):
            employee_name = clean_string(team_sheet.cell(row_index, 3).value)
            if not employee_name:
                continue
            employee_id = normalize_emp_code(team_sheet.cell(row_index, 1).value)
            employee = employees.get(employee_id or normalize_name(employee_name))
            if not employee:
                employee = EmployeeRecord(
                    employee_id=employee_id or normalize_name(employee_name),
                    name=employee_name,
                    designation=clean_string(team_sheet.cell(row_index, 7).value),
                    location=clean_string(team_sheet.cell(row_index, 8).value),
                )
                employees[employee.employee_id] = employee
            if month_name not in employee.month_order:
                employee.month_order.append(month_name)

            for column_index in range(13, team_sheet.max_column + 1):
                project_name = clean_string(team_sheet.cell(1, column_index).value)
                project_key = normalize_name(project_name)
                if not project_name or project_key == "total" or project_key not in project_meta:
                    continue
                share = parse_number(team_sheet.cell(row_index, column_index).value)
                if share == 0:
                    continue

                meta = project_meta.get(project_key, {})
                project_id = clean_string(meta.get("projectId")) or f"{source_sheet_name}-{column_index}"
                project = next((item for item in employee.projects if item.project_id == project_id), None)
                if not project:
                    closure = (closure_map or {}).get(project_id) or (closure_map or {}).get(normalize_name(project_name), {})
                    ytd = (ytd_map or {}).get(project_id) or (ytd_map or {}).get(normalize_name(project_name), {})
                    project = ProjectRecord(
                        project_id=project_id,
                        project_name=meta.get("projectName") or project_name,
                        project_value=parse_number(meta.get("projectValue")),
                        future_value=0.0,
                        closure_roles=list(closure.get("roles", [])),
                        sourcing_type=clean_string(closure.get("sourcingType")),
                        ytd_meta=dict(ytd),
                        source_sheet=source_sheet_name,
                    )
                    employee.projects.append(project)

                project.monthly_incentive[month_name] = share
                project.monthly_cf[month_name] = parse_number(meta.get("cashflow"))
                if parse_number(meta.get("projectValue")):
                    project.project_value = parse_number(meta.get("projectValue"))

    return list(employees.values())
