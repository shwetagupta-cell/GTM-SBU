from copy import copy
from pathlib import Path
import re

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = Path("/Users/shwetagupta/Documents/New project")
TARGET_WORKBOOK = Path("/Users/shwetagupta/Downloads/SME Incentives MIS.xlsx")
NEW_CLOSURE_WORKBOOK = Path("/Users/shwetagupta/Downloads/Feb & Mar SME DEAL CLOSURE MONTHLY DATA -2.xlsx")
OLD_CLOSURE_WORKBOOK = Path("/Users/shwetagupta/Downloads/SME DEAL CLOSURE DATA  (1).xlsx")
OUTPUT_DIR = ROOT / "outputs" / "feb-march-incentives"
OUTPUT_PATH = OUTPUT_DIR / "SME Incentives MIS - Feb March mapped.xlsx"

SOURCE_SHEET = "Feb & March26"
TEAM_SHEET = "Feb & March26 - Team Incentives"
TEAM_PROJECT_START_COL = 13  # M


ROLE_ROW_MAP = {
    ("Sourcing", "VT - Regional Head"): 21,
    ("Sourcing", "VT - BD Manager"): 22,
    ("Sourcing", "DM - DGM"): 23,
    ("Sourcing", "DM - Executive"): 24,
    ("Sourcing", "KAM - National Head"): 25,
    ("Sourcing", "KAM - Regional Head"): 26,
    ("Sourcing", "KAM - Sourcing Mgr"): 27,
    ("Closing", "BDM - Sales"): 30,
    ("Closing", "Sales Support Mgr"): 31,
    ("Closing", "Sales Support Mgr(Sanket)"): 31,
    ("Closing", "Sales Support Mgr(Sanket) "): 31,
    ("Closing", "MEP"): 32,
    ("Closing", "MEP "): 32,
    ("Closing", "MEP (Mazhar/Arif)"): 32,
    ("Closing", "MEP (Mazhar/Arif) "): 32,
    ("Pre Sales Design", "DGM / AGM Design"): 33,
    ("Pre Sales Design", "DGM/AGM Design"): 33,
    ("Pre Sales Design", "DGM/AGM  Design"): 33,
    ("Pre Sales Design", "Design Manager"): 34,
    ("Pre Sales Design", "Interior Designer"): 35,
    ("Pre Sales Design", "3d Artist"): 36,
    ("Pre Sales Design", "3d Artist "): 36,
    ("Pre Sales Design", "3D Artist"): 36,
    ("Pre Sales Design", "Developer / Modeller"): 37,
    ("Post Sales Design", "Design Manager"): 38,
    ("Post Sales Design", "Interior Designer"): 39,
    ("Post Sales Design", "3d Artist"): 40,
    ("Post Sales Design", "3d Artist "): 40,
    ("Post Sales Design", "3D Artist"): 40,
    ("Post Sales Design", "MeP Designer- Post Sales"): 41,
    ("Post Sales Design", "QS"): 42,
}


ALIASES = {
    "abu ansari": "Abu Sufiyan Irfan Ansari",
    "abu sufiyan": "Abu Sufiyan Irfan Ansari",
    "bharadwaaj": "Bharadwaaj Manchala",
    "bharadwaj": "Bharadwaaj Manchala",
    "charan": "Charanjeet Singh",
    "chandra kanth": "Chandrakanth Gupta K",
    "jitendra": "Jitendra Singh",
    "jeetendra": "Jeetendra Gupta",
    "mazhar": "Mazharali Barodawala",
    "neha": "Neha Titoria",
    "nikita bangar": "Nikita Rahul Bangar",
    "radhika deshpande": "Radhika Pranav Deshpande",
    "radhika despande": "Radhika Pranav Deshpande",
    "rohit": "Rohit Yadav",
    "rahul shirke": "Rahul Chandrakant Shirke",
    "samruddhi jadav": "Samruddhi Santosh Jadhav",
    "sufiyan ansari": "Abu Sufiyan Irfan Ansari",
    "sufiyaan ansari": "Abu Sufiyan Irfan Ansari",
    "sufiyan sheikh": "Mohammed Sufiyan Shaikh",
    "sufiyaan sheikh": "Mohammed Sufiyan Shaikh",
    "sufiyan sheihk": "Mohammed Sufiyan Shaikh",
    "tanvi chaurdary": "Tanvi Chowdhary",
    "utsav doshi": "Utsav Jayesh Doshi",
    "utsav rathod": "Utsav Nitin Rathod",
    "varun shetty": "Varun M Shetty",
    "zeba": "Zeba Baksh",
    "zeba baksh": "Zeba Baksh",
}


def normalize_text(value):
    return re.sub(r"[^a-z0-9]+", " ", str(value or "").lower()).strip()


def compact_space(value):
    return " ".join(str(value or "").replace("\n", " ").split())


def split_members(value):
    text = compact_space(value)
    if not text or text.lower() in {"na", "-", "?"}:
        return []
    return [
        part.strip()
        for part in re.split(r"\s*(?:\+|/|,|&|\n)\s*", text)
        if part.strip() and part.strip().lower() not in {"na", "-", "?"}
    ]


def member_tokens(value):
    return [token for token in normalize_text(value).split() if len(token) >= 2]


def canonical_function(function_name):
    normalized = normalize_text(function_name)
    if normalized == "sourcing":
        return "Sourcing"
    if normalized == "closing":
        return "Closing"
    if normalized.startswith("pre sales"):
        return "Pre Sales Design"
    if normalized.startswith("post sales"):
        return "Post Sales Design"
    return compact_space(function_name)


def build_employee_records(sheet):
    employees = []
    by_name = {}
    for row_index in range(2, sheet.max_row + 1):
        name = sheet.cell(row_index, 3).value
        if not name:
            continue
        record = {
            "row": row_index,
            "name": str(name).strip(),
            "name_norm": normalize_text(name),
            "tokens": member_tokens(name),
            "designation": compact_space(sheet.cell(row_index, 7).value),
            "department": compact_space(sheet.cell(row_index, 9).value),
            "pre_post": compact_space(sheet.cell(row_index, 12).value),
        }
        employees.append(record)
        by_name[record["name"]] = record
    return employees, by_name


def role_fit_bonus(record, role):
    function_name = role["function"]
    designation = role["designation"]
    designation_norm = normalize_text(designation)
    employee_designation = normalize_text(record["designation"])
    department = normalize_text(record["department"])
    pre_post = normalize_text(record["pre_post"])

    bonus = 0

    if function_name == "Pre Sales Design":
        if "design" in department and "pre sales" in pre_post:
            bonus += 5
        if "interior" in designation_norm and "interior" in employee_designation:
            bonus += 3
        if "3d" in designation_norm and "3d" in employee_designation:
            bonus += 3
        if "design manager" in designation_norm and "manager" in employee_designation:
            bonus += 3
        if "dgm" in designation_norm and ("general manager" in employee_designation or "agm" in employee_designation):
            bonus += 3

    if function_name == "Post Sales Design":
        if "design" in department and "post sales" in pre_post:
            bonus += 5
        if "interior" in designation_norm and "interior" in employee_designation:
            bonus += 3
        if "3d" in designation_norm and "3d" in employee_designation:
            bonus += 3
        if "design manager" in designation_norm and "manager" in employee_designation:
            bonus += 3

    if function_name == "Sourcing":
        if "key account" in employee_designation and "kam" in designation_norm:
            bonus += 5
        if "business development" in employee_designation and ("vt" in designation_norm or "bd manager" in designation_norm):
            bonus += 5

    if function_name == "Closing":
        if "business development" in employee_designation and "bdm" in designation_norm:
            bonus += 5
        if "operations" in department and "sales support" in designation_norm:
            bonus += 5
        if "mep" in employee_designation and "mep" in designation_norm:
            bonus += 5

    return bonus


def choose_alias(member_name, role):
    key = normalize_text(member_name)
    if key in ALIASES:
        return ALIASES[key]
    if key == "rahul":
        return "Rahul Chandrakant Shirke" if role["designation"] == "Interior Designer" else "Rahul Bajaj"
    if key == "nikita":
        return "Nikita Rachayya Hiremath" if role["function"] == "Post Sales Design" else "Nikita Rahul Bangar"
    if key == "shibin":
        return "Shibin M" if role["function"] == "Post Sales Design" else "Shibin P S"
    if key == "priya":
        return "Priya Khushal Deotale"
    if key == "rehan":
        return "Rehan Adil"
    if key == "shruti":
        return "Shruti Nargolkar"
    if key == "shiva":
        return "Shiva Prasad Goud Danthuri"
    if key == "sahith":
        return None
    return None


def match_member(member_name, role, employees, by_name):
    alias_name = choose_alias(member_name, role)
    if alias_name and alias_name in by_name:
        return by_name[alias_name]

    member_norm = normalize_text(member_name)
    member_parts = member_tokens(member_name)
    if not member_parts:
        return None

    best_record = None
    best_score = 0
    second_score = 0

    for record in employees:
        score = 0
        if record["name_norm"] == member_norm:
            score = 100
        else:
            exact_hits = 0
            partial_hits = 0
            for token in member_parts:
                if token in record["tokens"]:
                    exact_hits += 1
                elif any(candidate.startswith(token) or token.startswith(candidate) for candidate in record["tokens"]):
                    partial_hits += 1
                elif token in record["name_norm"]:
                    partial_hits += 1

            if exact_hits == len(member_parts):
                score += 40
            score += exact_hits * 20
            score += partial_hits * 8
            score += role_fit_bonus(record, role)

        if score > best_score:
            second_score = best_score
            best_score = score
            best_record = record
        elif score > second_score:
            second_score = score

    if best_record is None:
        return None
    if best_score < 20:
        return None
    if best_score - second_score <= 2 and best_score < 100:
        return None
    return best_record


def parse_closure_workbooks():
    sources = [
        (NEW_CLOSURE_WORKBOOK, ["Feb 26", "Mar 26"]),
        (OLD_CLOSURE_WORKBOOK, ["NOV 25", "DEC25", "Jan 26"]),
    ]
    projects = []

    for path, sheet_names in sources:
        workbook = load_workbook(path, data_only=False)
        for sheet_name in sheet_names:
            sheet = workbook[sheet_name]
            for column_index in range(3, sheet.max_column + 1):
                project_name = sheet.cell(2, column_index).value
                if not project_name:
                    continue
                project = {
                    "sheet": sheet_name,
                    "project_name": compact_space(project_name),
                    "project_id": compact_space(sheet.cell(3, column_index).value) or None,
                    "roles": [],
                }

                current_function = None
                for row_index in range(10, 31):
                    function_cell = sheet.cell(row_index, 1).value
                    designation = compact_space(sheet.cell(row_index, 2).value)
                    member_value = sheet.cell(row_index, column_index).value
                    members = split_members(member_value)

                    if function_cell:
                        current_function = canonical_function(function_cell)

                    if not designation or not members:
                        continue

                    project["roles"].append(
                        {
                            "function": current_function,
                            "designation": designation,
                            "members": members,
                        }
                    )

                projects.append(project)

    return projects


def find_closure_project(projects, project_name, project_id):
    project_id_norm = normalize_text(project_id)
    project_name_norm = normalize_text(project_name)

    for project in projects:
        if project_id_norm and normalize_text(project["project_id"]) == project_id_norm:
            return project

    for project in projects:
        if project_name_norm and normalize_text(project["project_name"]) == project_name_norm:
            return project

    return None


def build_active_project_list(sheet, closure_projects):
    active_projects = []

    for column_index in range(6, sheet.max_column + 1, 3):
        value_col = column_index + 1
        project_name = sheet.cell(2, value_col).value
        project_id = sheet.cell(3, value_col).value
        cashflow = sheet.cell(6, value_col).value

        try:
            numeric_cashflow = float(cashflow or 0)
        except (TypeError, ValueError):
            numeric_cashflow = 0

        if not project_name or numeric_cashflow == 0:
            continue

        closure_project = find_closure_project(closure_projects, project_name, project_id)
        active_projects.append(
            {
                "source_data_col": value_col,
                "project_name": compact_space(project_name),
                "project_id": compact_space(project_id) or None,
                "cashflow": numeric_cashflow,
                "closure_project": closure_project,
            }
        )

    return active_projects


def clear_existing_project_area(sheet, start_col, end_col):
    for row_index in range(1, sheet.max_row + 1):
        for column_index in range(start_col, end_col + 1):
            cell = sheet.cell(row_index, column_index)
            cell.value = None


def copy_project_column_style(sheet, source_col, target_col):
    sheet.column_dimensions[get_column_letter(target_col)].width = sheet.column_dimensions[get_column_letter(source_col)].width
    for row_index in range(1, sheet.max_row + 1):
        source_cell = sheet.cell(row_index, source_col)
        target_cell = sheet.cell(row_index, target_col)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)


def write_team_incentive_sheet(workbook):
    source_sheet = workbook[SOURCE_SHEET]
    team_sheet = workbook[TEAM_SHEET]

    employees, employees_by_name = build_employee_records(team_sheet)
    closure_projects = parse_closure_workbooks()
    active_projects = build_active_project_list(source_sheet, closure_projects)

    existing_last_col = team_sheet.max_column
    clear_existing_project_area(team_sheet, TEAM_PROJECT_START_COL, existing_last_col)

    template_col = 20 if existing_last_col >= 20 else TEAM_PROJECT_START_COL

    unmatched_members = set()
    project_formulas = {}

    for project_offset, project in enumerate(active_projects):
        column_index = TEAM_PROJECT_START_COL + project_offset
        if column_index > existing_last_col:
            copy_project_column_style(team_sheet, template_col, column_index)

        source_col_letter = get_column_letter(project["source_data_col"])
        team_sheet.cell(1, column_index).value = f"='{SOURCE_SHEET}'!{source_col_letter}2"

        formulas_by_row = {}
        closure_project = project["closure_project"]
        if closure_project:
            for role in closure_project["roles"]:
                source_row = ROLE_ROW_MAP.get((role["function"], role["designation"]))
                if source_row is None:
                    continue

                member_count = len(role["members"])
                if member_count == 0:
                    continue

                for member_name in role["members"]:
                    employee = match_member(member_name, role, employees, employees_by_name)
                    if employee is None:
                        unmatched_members.add(f"{project['project_name']} | {role['designation']} | {member_name}")
                        continue

                    source_ref = f"'{SOURCE_SHEET}'!{source_col_letter}{source_row}"
                    term = source_ref if member_count == 1 else f"{source_ref}/{member_count}"
                    formulas_by_row.setdefault(employee["row"], []).append(term)

        for row_index, terms in formulas_by_row.items():
            formula = "=" + "+".join(terms)
            team_sheet.cell(row_index, column_index).value = formula

        project_formulas[project["project_name"]] = len(formulas_by_row)

    workbook.calculation.forceFullCalc = True
    workbook.calculation.fullCalcOnLoad = True

    return active_projects, project_formulas, sorted(unmatched_members)


def main():
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    workbook = load_workbook(TARGET_WORKBOOK, data_only=False)
    active_projects, project_formulas, unmatched_members = write_team_incentive_sheet(workbook)
    workbook.save(OUTPUT_PATH)

    print(f"Saved workbook: {OUTPUT_PATH}")
    print(f"Active projects mapped: {len(active_projects)}")
    for project in active_projects:
        closure_sheet = project['closure_project']['sheet'] if project['closure_project'] else 'NO MATCH'
        print(
            f"- {project['project_name']} | CF={project['cashflow']} | source_col={get_column_letter(project['source_data_col'])} | "
            f"closure={closure_sheet} | employee_rows={project_formulas.get(project['project_name'], 0)}"
        )

    print(f"Unmatched closure names: {len(unmatched_members)}")
    for item in unmatched_members[:60]:
        print(f"  {item}")


if __name__ == "__main__":
    main()
